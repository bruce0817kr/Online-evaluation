"""
평가 시스템 API 엔드포인트
"""
from fastapi import APIRouter, Depends, HTTPException, status
from typing import List, Optional
from datetime import datetime
import logging
from bson import ObjectId

from models import (
    Evaluation, EvaluationCreate, EvaluationCriterion,
    ScoreSubmission, ScoreSubmissionCreate, User
)
from security import (
    check_admin_or_secretary, check_evaluator_role, 
    check_evaluator_assignment, get_current_user,
    db
)

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/evaluations", tags=["evaluations"])

@router.post("/", status_code=status.HTTP_201_CREATED)
async def create_evaluation(
    evaluation_data: EvaluationCreate,
    current_user: User = Depends(check_admin_or_secretary)
) -> dict:
    """
    새로운 평가 프로젝트 생성 (관리자/간사 전용)
    """
    if db is None:
        raise HTTPException(status_code=500, detail="Database not initialized")
    
    try:
        # 평가 데이터 구성
        new_eval = {
            "title": evaluation_data.title,
            "description": evaluation_data.description or "",
            "companies": evaluation_data.companies,
            "criteria": [criterion.dict() for criterion in evaluation_data.criteria],
            "status": evaluation_data.status,
            "created_by": current_user.id,
            "assigned_evaluators": [],
            "created_at": datetime.utcnow(),
            "updated_at": datetime.utcnow()
        }
        
        result = await db.evaluations.insert_one(new_eval)
        
        logger.info(f"Evaluation created by {current_user.login_id}: {evaluation_data.title}")
        
        return {
            "evaluation_id": str(result.inserted_id),
            "status": "created",
            "message": "평가가 성공적으로 생성되었습니다."
        }
        
    except Exception as e:
        logger.error(f"Error creating evaluation: {e}")
        raise HTTPException(status_code=500, detail="평가 생성 중 오류가 발생했습니다.")

@router.get("/")
async def list_evaluations(
    current_user: User = Depends(get_current_user),
    assigned_to_me: bool = False
) -> List[dict]:
    """
    평가 목록 조회
    - 관리자/간사: 모든 평가 조회 가능
    - 평가위원: assigned_to_me=True일 때 자신에게 배정된 평가만 조회
    """
    if db is None:
        raise HTTPException(status_code=500, detail="Database not initialized")
    
    try:
        # 쿼리 조건 설정
        query = {}
        
        if current_user.role == "evaluator" and assigned_to_me:
            # 평가위원이 자신에게 배정된 평가만 조회
            query["assigned_evaluators"] = current_user.id
        elif current_user.role == "evaluator" and not assigned_to_me:
            # 평가위원이 전체 목록을 요청하면 권한 없음
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="평가위원은 배정된 평가만 조회할 수 있습니다."
            )
        
        # 평가 목록 조회
        cursor = db.evaluations.find(query)
        evaluations = []
        
        async for eval_doc in cursor:
            eval_dict = {
                "id": str(eval_doc["_id"]),
                "title": eval_doc["title"],
                "description": eval_doc["description"],
                "status": eval_doc["status"],
                "companies_count": len(eval_doc.get("companies", [])),
                "criteria_count": len(eval_doc.get("criteria", [])),
                "assigned_evaluators_count": len(eval_doc.get("assigned_evaluators", [])),
                "created_by": eval_doc["created_by"],
                "created_at": eval_doc["created_at"].isoformat()
            }
            evaluations.append(eval_dict)
        
        return evaluations
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error listing evaluations: {e}")
        raise HTTPException(status_code=500, detail="평가 목록 조회 중 오류가 발생했습니다.")

@router.get("/{evaluation_id}")
async def get_evaluation(
    evaluation_id: str,
    current_user: User = Depends(get_current_user)
) -> dict:
    """
    특정 평가 상세 조회
    """
    if db is None:
        raise HTTPException(status_code=500, detail="Database not initialized")
    
    try:
        # ObjectId 유효성 검사
        if not ObjectId.is_valid(evaluation_id):
            raise HTTPException(status_code=400, detail="Invalid evaluation ID")
        
        eval_doc = await db.evaluations.find_one({"_id": ObjectId(evaluation_id)})
        if not eval_doc:
            raise HTTPException(status_code=404, detail="평가를 찾을 수 없습니다.")
        
        # 권한 확인 (평가위원은 배정된 평가만 조회 가능)
        if current_user.role == "evaluator":
            assigned_evaluators = eval_doc.get("assigned_evaluators", [])
            if current_user.id not in assigned_evaluators:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail="이 평가에 대한 접근 권한이 없습니다."
                )
        
        # 응답 데이터 구성
        response = {
            "id": str(eval_doc["_id"]),
            "title": eval_doc["title"],
            "description": eval_doc["description"],
            "companies": eval_doc["companies"],
            "criteria": eval_doc["criteria"],
            "status": eval_doc["status"],
            "assigned_evaluators": eval_doc.get("assigned_evaluators", []),
            "created_by": eval_doc["created_by"],
            "created_at": eval_doc["created_at"].isoformat(),
            "updated_at": eval_doc["updated_at"].isoformat()
        }
        
        return response
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting evaluation {evaluation_id}: {e}")
        raise HTTPException(status_code=500, detail="평가 조회 중 오류가 발생했습니다.")

@router.post("/{evaluation_id}/assign")
async def assign_evaluators(
    evaluation_id: str,
    payload: dict,
    current_user: User = Depends(check_admin_or_secretary)
) -> dict:
    """
    평가에 평가위원 배정 (관리자/간사 전용)
    """
    if db is None:
        raise HTTPException(status_code=500, detail="Database not initialized")
    
    try:
        evaluator_ids = payload.get("evaluator_ids", [])
        
        # ObjectId 유효성 검사
        if not ObjectId.is_valid(evaluation_id):
            raise HTTPException(status_code=400, detail="Invalid evaluation ID")
        
        # 평가 존재 여부 확인
        evaluation = await db.evaluations.find_one({"_id": ObjectId(evaluation_id)})
        if not evaluation:
            raise HTTPException(status_code=404, detail="평가를 찾을 수 없습니다.")
        
        # 평가위원 ID들이 실제 평가위원 역할인지 확인
        if evaluator_ids:
            evaluators = await db.users.find({
                "_id": {"$in": [ObjectId(eid) for eid in evaluator_ids if ObjectId.is_valid(eid)]},
                "role": "evaluator",
                "is_active": True
            }).to_list(length=None)
            
            found_ids = [str(evaluator["_id"]) for evaluator in evaluators]
            invalid_ids = set(evaluator_ids) - set(found_ids)
            
            if invalid_ids:
                raise HTTPException(
                    status_code=400, 
                    detail=f"유효하지 않은 평가위원 ID: {list(invalid_ids)}"
                )
        
        # 평가위원 배정 업데이트
        await db.evaluations.update_one(
            {"_id": ObjectId(evaluation_id)},
            {
                "$set": {
                    "assigned_evaluators": evaluator_ids,
                    "updated_at": datetime.utcnow()
                }
            }
        )
        
        logger.info(f"Evaluators assigned to evaluation {evaluation_id} by {current_user.login_id}")
        
        return {
            "assigned_count": len(evaluator_ids),
            "message": f"{len(evaluator_ids)}명의 평가위원이 배정되었습니다."
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error assigning evaluators to evaluation {evaluation_id}: {e}")
        raise HTTPException(status_code=500, detail="평가위원 배정 중 오류가 발생했습니다.")

@router.post("/{evaluation_id}/submit")
async def submit_scores(
    evaluation_id: str,
    submission: ScoreSubmissionCreate,
    current_user: User = Depends(check_evaluator_assignment)
) -> dict:
    """
    평가 점수 제출 (배정된 평가위원 전용)
    """
    if db is None:
        raise HTTPException(status_code=500, detail="Database not initialized")
    
    try:
        # ObjectId 유효성 검사
        if not ObjectId.is_valid(evaluation_id):
            raise HTTPException(status_code=400, detail="Invalid evaluation ID")
        
        # 평가 정보 조회
        eval_doc = await db.evaluations.find_one({"_id": ObjectId(evaluation_id)})
        if not eval_doc:
            raise HTTPException(status_code=404, detail="평가를 찾을 수 없습니다.")
        
        # 중복 제출 확인
        existing_submission = await db.submissions.find_one({
            "evaluation_id": evaluation_id,
            "company_id": submission.companyId,
            "evaluator_id": current_user.id
        })
        
        if existing_submission:
            raise HTTPException(
                status_code=400, 
                detail="이미 해당 기업에 대한 평가를 제출했습니다."
            )
        
        # 점수 검증 및 총점 계산
        total_score = 0
        criteria_dict = {crit["name"]: crit for crit in eval_doc["criteria"]}
        
        for criterion_name, score in submission.scores.items():
            if criterion_name not in criteria_dict:
                raise HTTPException(
                    status_code=400, 
                    detail=f"유효하지 않은 평가 항목: {criterion_name}"
                )
            
            max_score = criteria_dict[criterion_name]["max_score"]
            if score < 0 or score > max_score:
                raise HTTPException(
                    status_code=400,
                    detail=f"{criterion_name} 점수는 0-{max_score} 범위여야 합니다."
                )
            
            total_score += score
        
        # 제출 데이터 저장
        submission_data = {
            "evaluation_id": evaluation_id,
            "company_id": submission.companyId,
            "evaluator_id": current_user.id,
            "scores": submission.scores,
            "comment": submission.comment,
            "total_score": total_score,
            "submitted_at": datetime.utcnow()
        }
        
        result = await db.submissions.insert_one(submission_data)
        
        logger.info(f"Score submitted by {current_user.login_id} for evaluation {evaluation_id}")
        
        return {
            "submission_id": str(result.inserted_id),
            "total_score": total_score,
            "message": "평가가 성공적으로 제출되었습니다."
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error submitting scores for evaluation {evaluation_id}: {e}")
        raise HTTPException(status_code=500, detail="점수 제출 중 오류가 발생했습니다.")

@router.get("/{evaluation_id}/progress")
async def get_evaluation_progress(
    evaluation_id: str,
    current_user: User = Depends(check_admin_or_secretary)
) -> dict:
    """
    평가 진행 상황 조회 (관리자/간사 전용)
    """
    if db is None:
        raise HTTPException(status_code=500, detail="Database not initialized")
    
    try:
        # ObjectId 유효성 검사
        if not ObjectId.is_valid(evaluation_id):
            raise HTTPException(status_code=400, detail="Invalid evaluation ID")
        
        # 평가 정보 조회
        eval_doc = await db.evaluations.find_one({"_id": ObjectId(evaluation_id)})
        if not eval_doc:
            raise HTTPException(status_code=404, detail="평가를 찾을 수 없습니다.")
        
        companies = eval_doc.get("companies", [])
        assigned_evaluators = eval_doc.get("assigned_evaluators", [])
        
        # 제출된 점수 조회
        submissions = await db.submissions.find({
            "evaluation_id": evaluation_id
        }).to_list(length=None)
        
        # 진행률 계산
        total_expected = len(companies) * len(assigned_evaluators)
        completed_submissions = len(submissions)
        progress_percentage = (completed_submissions / total_expected * 100) if total_expected > 0 else 0
        
        # 평가위원별 진행 상황
        evaluator_progress = {}
        for evaluator_id in assigned_evaluators:
            evaluator_submissions = [s for s in submissions if s["evaluator_id"] == evaluator_id]
            evaluator_progress[evaluator_id] = {
                "completed": len(evaluator_submissions),
                "total": len(companies),
                "percentage": (len(evaluator_submissions) / len(companies) * 100) if companies else 0
            }
        
        return {
            "evaluation_id": evaluation_id,
            "total_companies": len(companies),
            "total_evaluators": len(assigned_evaluators),
            "total_expected_submissions": total_expected,
            "completed_submissions": completed_submissions,
            "progress_percentage": round(progress_percentage, 2),
            "evaluator_progress": evaluator_progress
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting evaluation progress {evaluation_id}: {e}")
        raise HTTPException(status_code=500, detail="진행 상황 조회 중 오류가 발생했습니다.")

# Excel 대량 업로드 관련 엔드포인트
from fastapi import File, UploadFile
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from fastapi.responses import StreamingResponse

@router.post("/bulk-upload-excel")
async def bulk_upload_evaluations_excel(
    file: UploadFile = File(...),
    current_user: User = Depends(check_admin_or_secretary)
):
    """Excel 파일을 통한 평가 대량 생성"""
    if db is None:
        raise HTTPException(status_code=500, detail="Database not initialized")
    
    # 파일 타입 검증
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Excel 파일만 지원됩니다 (.xlsx, .xls)")
    
    try:
        # Excel 파일 읽기
        content = await file.read()
        excel_data = pd.ExcelFile(io.BytesIO(content))
        
        # 필수 시트 확인
        required_sheets = ['Evaluation_Info', 'Companies', 'Criteria']
        missing_sheets = [sheet for sheet in required_sheets if sheet not in excel_data.sheet_names]
        if missing_sheets:
            raise HTTPException(
                status_code=400, 
                detail=f"필수 시트가 없습니다: {', '.join(missing_sheets)}"
            )
        
        # 데이터 파싱
        results = {
            "success_count": 0,
            "failed_count": 0,
            "errors": [],
            "created_evaluation_ids": []
        }
        
        # 평가 정보 시트 파싱
        eval_info_df = pd.read_excel(io.BytesIO(content), sheet_name='Evaluation_Info', header=None)
        eval_info = dict(zip(eval_info_df.iloc[:, 0], eval_info_df.iloc[:, 1]))
        
        # 기업 목록 시트 파싱
        companies_df = pd.read_excel(io.BytesIO(content), sheet_name='Companies')
        companies = companies_df['company_name'].dropna().tolist()
        
        # 평가 기준 시트 파싱
        criteria_df = pd.read_excel(io.BytesIO(content), sheet_name='Criteria')
        criteria = []
        
        for index, row in criteria_df.iterrows():
            try:
                criterion = {
                    "name": str(row['name']),
                    "max_score": int(row['max_score']),
                    "bonus": bool(row.get('bonus', False)),
                    "description": str(row.get('description', ''))
                }
                criteria.append(criterion)
            except Exception as e:
                results["errors"].append({
                    "row": index + 2,
                    "sheet": "Criteria",
                    "error": f"기준 파싱 오류: {str(e)}",
                    "value": str(row.to_dict())
                })
                results["failed_count"] += 1
        
        # 데이터 유효성 검증
        if not eval_info.get('title'):
            raise HTTPException(status_code=400, detail="평가 제목이 필요합니다")
        
        if len(companies) == 0:
            raise HTTPException(status_code=400, detail="최소 1개 이상의 기업이 필요합니다")
        
        if len(criteria) == 0:
            raise HTTPException(status_code=400, detail="최소 1개 이상의 평가 기준이 필요합니다")
        
        # 평가 생성
        try:
            new_eval = {
                "title": eval_info['title'],
                "description": eval_info.get('description', ''),
                "companies": companies,
                "criteria": criteria,
                "status": eval_info.get('status', 'draft'),
                "created_by": current_user.id,
                "assigned_evaluators": [],
                "created_at": datetime.utcnow(),
                "updated_at": datetime.utcnow()
            }
            
            result = await db.evaluations.insert_one(new_eval)
            results["created_evaluation_ids"].append(str(result.inserted_id))
            results["success_count"] = 1
            
            logger.info(f"Bulk evaluation created by {current_user.login_id}: {eval_info['title']}")
            
        except Exception as e:
            results["errors"].append({
                "sheet": "General",
                "error": f"평가 생성 실패: {str(e)}"
            })
            results["failed_count"] += 1
        
        return results
        
    except pd.errors.ParserError as e:
        raise HTTPException(status_code=400, detail=f"Excel 파일 파싱 오류: {str(e)}")
    except Exception as e:
        logger.error(f"Excel bulk upload error: {e}")
        raise HTTPException(status_code=500, detail=f"Excel 업로드 중 오류가 발생했습니다: {str(e)}")

@router.get("/download-template")
async def download_excel_template():
    """Excel 템플릿 파일 다운로드"""
    try:
        # 새 워크북 생성
        wb = Workbook()
        
        # 평가 정보 시트
        ws1 = wb.active
        ws1.title = "Evaluation_Info"
        ws1.append(["Field", "Value"])
        ws1.append(["title", "2025년 상반기 혁신기업 평가"])
        ws1.append(["description", "평가 설명을 입력하세요"])
        ws1.append(["status", "draft"])
        
        # 기업 목록 시트
        ws2 = wb.create_sheet("Companies")
        ws2.append(["company_name"])
        ws2.append(["삼성전자"])
        ws2.append(["LG전자"])
        ws2.append(["현대자동차"])
        
        # 평가 기준 시트
        ws3 = wb.create_sheet("Criteria")
        ws3.append(["name", "max_score", "bonus", "description"])
        ws3.append(["기술성", 100, False, "기술 수준 및 혁신성"])
        ws3.append(["사업성", 100, False, "사업 가능성 및 수익성"])
        ws3.append(["가점항목", 20, True, "추가 가점 요소"])
        
        # 메모리에서 Excel 파일 생성
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 응답 헤더 설정
        headers = {
            'Content-Disposition': 'attachment; filename="evaluation_template.xlsx"',
            'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }
        
        return StreamingResponse(
            io.BytesIO(output.getvalue()), 
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers
        )
        
    except Exception as e:
        logger.error(f"Template download error: {e}")
        raise HTTPException(status_code=500, detail="템플릿 다운로드 중 오류가 발생했습니다")

@router.post("/validate-excel")
async def validate_excel(
    file: UploadFile = File(...),
    current_user: User = Depends(check_admin_or_secretary)
):
    """Excel 파일 유효성 검증 (업로드 전 미리보기)"""
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Excel 파일만 지원됩니다")
    
    try:
        content = await file.read()
        excel_data = pd.ExcelFile(io.BytesIO(content))
        
        # 시트 존재 여부 확인
        required_sheets = ['Evaluation_Info', 'Companies', 'Criteria']
        validation_result = {
            "valid": True,
            "errors": [],
            "warnings": [],
            "preview": {}
        }
        
        # 필수 시트 확인
        missing_sheets = [sheet for sheet in required_sheets if sheet not in excel_data.sheet_names]
        if missing_sheets:
            validation_result["valid"] = False
            validation_result["errors"].append(f"필수 시트가 없습니다: {', '.join(missing_sheets)}")
            return validation_result
        
        # 각 시트 데이터 미리보기
        try:
            # 평가 정보
            eval_info_df = pd.read_excel(io.BytesIO(content), sheet_name='Evaluation_Info', header=None)
            eval_info = dict(zip(eval_info_df.iloc[:, 0], eval_info_df.iloc[:, 1]))
            validation_result["preview"]["evaluation_info"] = eval_info
            
            if not eval_info.get('title'):
                validation_result["errors"].append("평가 제목이 필요합니다")
                validation_result["valid"] = False
            
        except Exception as e:
            validation_result["errors"].append(f"평가 정보 시트 오류: {str(e)}")
            validation_result["valid"] = False
        
        try:
            # 기업 목록
            companies_df = pd.read_excel(io.BytesIO(content), sheet_name='Companies')
            companies = companies_df['company_name'].dropna().tolist()
            validation_result["preview"]["companies"] = companies[:10]  # 처음 10개만 미리보기
            validation_result["preview"]["total_companies"] = len(companies)
            
            if len(companies) == 0:
                validation_result["errors"].append("최소 1개 이상의 기업이 필요합니다")
                validation_result["valid"] = False
            
        except Exception as e:
            validation_result["errors"].append(f"기업 목록 시트 오류: {str(e)}")
            validation_result["valid"] = False
        
        try:
            # 평가 기준
            criteria_df = pd.read_excel(io.BytesIO(content), sheet_name='Criteria')
            criteria_preview = []
            
            for index, row in criteria_df.iterrows():
                if index >= 5:  # 처음 5개만 미리보기
                    break
                criteria_preview.append({
                    "name": str(row['name']),
                    "max_score": int(row['max_score']),
                    "bonus": bool(row.get('bonus', False)),
                    "description": str(row.get('description', ''))
                })
            
            validation_result["preview"]["criteria"] = criteria_preview
            validation_result["preview"]["total_criteria"] = len(criteria_df)
            
            if len(criteria_df) == 0:
                validation_result["errors"].append("최소 1개 이상의 평가 기준이 필요합니다")
                validation_result["valid"] = False
            
        except Exception as e:
            validation_result["errors"].append(f"평가 기준 시트 오류: {str(e)}")
            validation_result["valid"] = False
        
        return validation_result
        
    except Exception as e:
        logger.error(f"Excel validation error: {e}")
        raise HTTPException(status_code=500, detail=f"Excel 유효성 검증 중 오류가 발생했습니다: {str(e)}")
