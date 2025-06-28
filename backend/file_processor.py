"""
고성능 파일 처리 모듈
대용량 파일 업로드, 처리, 변환을 위한 최적화된 서비스
"""

import os
import asyncio
import aiofiles
from pathlib import Path
from typing import Optional, Dict, Any, List, BinaryIO
import logging
import hashlib
import mimetypes
from datetime import datetime
import tempfile
import shutil

# 파일 처리 라이브러리들
try:
    import PyPDF2
    import pdfplumber
    from docx import Document as DocxDocument
    from openpyxl import load_workbook
    import pandas as pd
    HAS_FILE_LIBS = True
except ImportError:
    HAS_FILE_LIBS = False
    print("⚠️ 파일 처리 라이브러리가 설치되지 않았습니다. pip install PyPDF2 pdfplumber python-docx openpyxl pandas 를 실행하세요.")

# 이미지 처리 라이브러리
try:
    from PIL import Image
    import pytesseract
    HAS_OCR = True
except ImportError:
    HAS_OCR = False
    print("⚠️ OCR 라이브러리가 설치되지 않았습니다. pip install Pillow pytesseract 를 실행하세요.")

logger = logging.getLogger(__name__)

class FileProcessor:
    """고성능 파일 처리 클래스"""
    
    def __init__(self, upload_dir: str = "./uploads", temp_dir: str = "./temp"):
        self.upload_dir = Path(upload_dir)
        self.temp_dir = Path(temp_dir)
        
        # 디렉토리 생성
        self.upload_dir.mkdir(exist_ok=True)
        self.temp_dir.mkdir(exist_ok=True)
        
        # 지원하는 파일 형식
        self.supported_formats = {
            'documents': ['.pdf', '.docx', '.doc', '.txt'],
            'spreadsheets': ['.xlsx', '.xls', '.csv'],
            'images': ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff'],
            'archives': ['.zip', '.rar', '.7z']
        }
        
        # 파일 크기 제한 (바이트)
        self.max_file_sizes = {
            'default': 100 * 1024 * 1024,  # 100MB
            'pdf': 200 * 1024 * 1024,     # 200MB
            'image': 50 * 1024 * 1024,    # 50MB
            'archive': 500 * 1024 * 1024  # 500MB
        }
    
    async def process_uploaded_file(
        self, 
        file_data: BinaryIO, 
        filename: str, 
        content_type: str,
        company_id: str,
        uploaded_by: str
    ) -> Dict[str, Any]:
        """업로드된 파일 처리"""
        try:
            # 파일 정보 분석
            file_info = await self._analyze_file(file_data, filename, content_type)
            
            # 보안 검사
            security_check = await self._security_scan(file_data, file_info)
            if not security_check['safe']:
                raise ValueError(f"보안 검사 실패: {security_check['reason']}")
            
            # 파일 저장
            file_path = await self._save_file_async(file_data, filename, company_id)
            
            # 텍스트 추출 (필요한 경우)
            extracted_text = None
            if file_info['category'] in ['documents', 'images']:
                extracted_text = await self._extract_text_async(file_path, file_info)
            
            # 메타데이터 생성
            metadata = {
                'id': self._generate_file_id(),
                'filename': filename,
                'original_filename': filename,
                'file_path': str(file_path),
                'file_size': file_info['size'],
                'file_type': content_type,
                'file_hash': file_info['hash'],
                'company_id': company_id,
                'uploaded_by': uploaded_by,
                'uploaded_at': datetime.utcnow(),
                'processed_at': datetime.utcnow(),
                'category': file_info['category'],
                'extracted_text': extracted_text,
                'page_count': file_info.get('page_count'),
                'processing_status': 'completed'
            }
            
            logger.info(f"파일 처리 완료: {filename}", extra={
                'file_size': file_info['size'],
                'category': file_info['category'],
                'company_id': company_id
            })
            
            return metadata
            
        except Exception as e:
            logger.error(f"파일 처리 오류: {e}", extra={
                'filename': filename,
                'company_id': company_id,
                'error': str(e)
            })
            raise
    
    async def _analyze_file(self, file_data: BinaryIO, filename: str, content_type: str) -> Dict[str, Any]:
        """파일 분석"""
        # 파일 크기 계산
        file_data.seek(0, 2)  # 파일 끝으로 이동
        file_size = file_data.tell()
        file_data.seek(0)  # 파일 시작으로 되돌리기
        
        # 파일 확장자 및 카테고리 결정
        file_ext = Path(filename).suffix.lower()
        category = self._get_file_category(file_ext)
        
        # 파일 크기 제한 확인
        max_size = self.max_file_sizes.get(category, self.max_file_sizes['default'])
        if file_size > max_size:
            raise ValueError(f"파일 크기가 제한을 초과했습니다. 최대 {max_size // (1024*1024)}MB")
        
        # 파일 해시 계산
        file_hash = await self._calculate_file_hash(file_data)
        
        # PDF 파일의 경우 페이지 수 계산
        page_count = None
        if file_ext == '.pdf' and HAS_FILE_LIBS:
            try:
                page_count = await self._count_pdf_pages(file_data)
            except:
                pass
        
        return {
            'size': file_size,
            'extension': file_ext,
            'category': category,
            'hash': file_hash,
            'mime_type': content_type,
            'page_count': page_count
        }
    
    async def _security_scan(self, file_data: BinaryIO, file_info: Dict[str, Any]) -> Dict[str, Any]:
        """기본 보안 검사"""
        # 파일 시그니처 검사
        file_data.seek(0)
        file_header = file_data.read(16)
        file_data.seek(0)
        
        # 악성 파일 시그니처 검사 (간단한 구현)
        malicious_signatures = [
            b'\x4d\x5a',  # PE 실행 파일
            b'\x7f\x45\x4c\x46',  # ELF 실행 파일
        ]
        
        for signature in malicious_signatures:
            if file_header.startswith(signature):
                return {
                    'safe': False,
                    'reason': '실행 파일은 업로드할 수 없습니다.'
                }
        
        # 파일 확장자와 실제 내용 일치성 검사
        expected_mime = mimetypes.guess_type(f"file{file_info['extension']}")[0]
        if expected_mime and expected_mime != file_info['mime_type']:
            logger.warning(f"MIME 타입 불일치: {expected_mime} vs {file_info['mime_type']}")
        
        return {'safe': True, 'reason': None}
    
    async def _save_file_async(self, file_data: BinaryIO, filename: str, company_id: str) -> Path:
        """비동기 파일 저장"""
        # 안전한 파일명 생성
        safe_filename = self._sanitize_filename(filename)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        unique_filename = f"{company_id}_{timestamp}_{safe_filename}"
        
        file_path = self.upload_dir / unique_filename
        
        # 비동기 파일 쓰기
        file_data.seek(0)
        async with aiofiles.open(file_path, 'wb') as f:
            chunk_size = 8192
            while True:
                chunk = file_data.read(chunk_size)
                if not chunk:
                    break
                await f.write(chunk)
        
        return file_path
    
    async def _extract_text_async(self, file_path: Path, file_info: Dict[str, Any]) -> Optional[str]:
        """비동기 텍스트 추출"""
        if not HAS_FILE_LIBS:
            return None
        
        try:
            extension = file_info['extension']
            
            if extension == '.pdf':
                return await self._extract_pdf_text(file_path)
            elif extension == '.docx':
                return await self._extract_docx_text(file_path)
            elif extension in ['.xlsx', '.xls']:
                return await self._extract_excel_text(file_path)
            elif extension == '.txt':
                return await self._extract_txt_text(file_path)
            elif extension in ['.jpg', '.jpeg', '.png'] and HAS_OCR:
                return await self._extract_image_text(file_path)
            
        except Exception as e:
            logger.error(f"텍스트 추출 오류: {e}", extra={'file_path': str(file_path)})
        
        return None
    
    async def _extract_pdf_text(self, file_path: Path) -> str:
        """PDF 텍스트 추출"""
        def extract_sync():
            try:
                with pdfplumber.open(file_path) as pdf:
                    text_parts = []
                    for page in pdf.pages:
                        page_text = page.extract_text()
                        if page_text:
                            text_parts.append(page_text)
                    return '\n'.join(text_parts)
            except:
                # Fallback to PyPDF2
                with open(file_path, 'rb') as file:
                    pdf_reader = PyPDF2.PdfReader(file)
                    text_parts = []
                    for page in pdf_reader.pages:
                        text_parts.append(page.extract_text())
                    return '\n'.join(text_parts)
        
        # 별도 스레드에서 실행
        loop = asyncio.get_event_loop()
        return await loop.run_in_executor(None, extract_sync)
    
    async def _extract_docx_text(self, file_path: Path) -> str:
        """DOCX 텍스트 추출"""
        def extract_sync():
            doc = DocxDocument(file_path)
            text_parts = []
            for paragraph in doc.paragraphs:
                text_parts.append(paragraph.text)
            return '\n'.join(text_parts)
        
        loop = asyncio.get_event_loop()
        return await loop.run_in_executor(None, extract_sync)
    
    async def _extract_excel_text(self, file_path: Path) -> str:
        """Excel 텍스트 추출"""
        def extract_sync():
            try:
                df = pd.read_excel(file_path)
                return df.to_string()
            except:
                # Fallback to openpyxl
                wb = load_workbook(file_path, data_only=True)
                text_parts = []
                for sheet in wb.worksheets:
                    for row in sheet.iter_rows(values_only=True):
                        row_text = [str(cell) for cell in row if cell is not None]
                        if row_text:
                            text_parts.append(' '.join(row_text))
                return '\n'.join(text_parts)
        
        loop = asyncio.get_event_loop()
        return await loop.run_in_executor(None, extract_sync)
    
    async def _extract_txt_text(self, file_path: Path) -> str:
        """텍스트 파일 읽기"""
        async with aiofiles.open(file_path, 'r', encoding='utf-8') as f:
            return await f.read()
    
    async def _extract_image_text(self, file_path: Path) -> str:
        """이미지 OCR 텍스트 추출"""
        def extract_sync():
            image = Image.open(file_path)
            return pytesseract.image_to_string(image, lang='kor+eng')
        
        loop = asyncio.get_event_loop()
        return await loop.run_in_executor(None, extract_sync)
    
    async def _calculate_file_hash(self, file_data: BinaryIO) -> str:
        """파일 해시 계산"""
        hash_sha256 = hashlib.sha256()
        file_data.seek(0)
        
        while True:
            chunk = file_data.read(8192)
            if not chunk:
                break
            hash_sha256.update(chunk)
        
        file_data.seek(0)
        return hash_sha256.hexdigest()
    
    async def _count_pdf_pages(self, file_data: BinaryIO) -> int:
        """PDF 페이지 수 계산"""
        def count_sync():
            try:
                pdf_reader = PyPDF2.PdfReader(file_data)
                return len(pdf_reader.pages)
            except:
                return 0
        
        file_data.seek(0)
        loop = asyncio.get_event_loop()
        page_count = await loop.run_in_executor(None, count_sync)
        file_data.seek(0)
        return page_count
    
    def _get_file_category(self, extension: str) -> str:
        """파일 카테고리 결정"""
        for category, extensions in self.supported_formats.items():
            if extension in extensions:
                return category
        return 'other'
    
    def _sanitize_filename(self, filename: str) -> str:
        """파일명 정리"""
        import re
        # 위험한 문자 제거
        safe_name = re.sub(r'[^\w\-_\.]', '_', filename)
        # 연속된 언더스코어 제거
        safe_name = re.sub(r'_+', '_', safe_name)
        return safe_name[:100]  # 길이 제한
    
    def _generate_file_id(self) -> str:
        """고유 파일 ID 생성"""
        import uuid
        return str(uuid.uuid4())
    
    async def optimize_file(self, file_path: Path, file_type: str) -> Optional[Path]:
        """파일 최적화 (압축, 크기 조정 등)"""
        try:
            if file_type == 'application/pdf':
                return await self._optimize_pdf(file_path)
            elif file_type.startswith('image/'):
                return await self._optimize_image(file_path)
            
        except Exception as e:
            logger.error(f"파일 최적화 오류: {e}")
        
        return None
    
    async def _optimize_pdf(self, file_path: Path) -> Path:
        """PDF 최적화"""
        # 간단한 구현 - 실제로는 더 복잡한 압축 알고리즘 사용
        optimized_path = file_path.with_suffix('.optimized.pdf')
        shutil.copy2(file_path, optimized_path)
        return optimized_path
    
    async def _optimize_image(self, file_path: Path) -> Path:
        """이미지 최적화"""
        if not HAS_OCR:
            return file_path
        
        def optimize_sync():
            image = Image.open(file_path)
            
            # 큰 이미지는 리사이즈
            max_size = (1920, 1080)
            if image.size[0] > max_size[0] or image.size[1] > max_size[1]:
                image.thumbnail(max_size, Image.Resampling.LANCZOS)
            
            # 압축률 높여서 저장
            optimized_path = file_path.with_suffix('.optimized.jpg')
            image.save(optimized_path, 'JPEG', quality=85, optimize=True)
            return optimized_path
        
        loop = asyncio.get_event_loop()
        return await loop.run_in_executor(None, optimize_sync)
    
    async def cleanup_temp_files(self, max_age_hours: int = 24):
        """임시 파일 정리"""
        try:
            current_time = datetime.now()
            for file_path in self.temp_dir.iterdir():
                if file_path.is_file():
                    file_age = current_time - datetime.fromtimestamp(file_path.stat().st_mtime)
                    if file_age.total_seconds() > max_age_hours * 3600:
                        file_path.unlink()
                        logger.info(f"임시 파일 삭제: {file_path}")
        except Exception as e:
            logger.error(f"임시 파일 정리 오류: {e}")

# 전역 파일 프로세서 인스턴스
file_processor = FileProcessor()