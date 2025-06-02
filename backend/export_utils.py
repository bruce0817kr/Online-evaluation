"""
종합평가서 추출 유틸리티 모듈
PDF와 Excel 형태로 평가 데이터를 추출하는 기능을 제공합니다.
"""

import io
import os
from datetime import datetime
from typing import List, Dict, Any, Optional
import pandas as pd
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import xlsxwriter
from pathlib import Path

# 한글 폰트 설정 (시스템에 설치된 폰트 사용)
try:
    # Windows 한글 폰트
    font_path = "C:/Windows/Fonts/malgun.ttf"
    if os.path.exists(font_path):
        pdfmetrics.registerFont(TTFont('Korean', font_path))
    else:
        # Linux/Mac 대체 폰트 (설치되어 있을 경우)
        alt_fonts = [
            "/usr/share/fonts/truetype/nanum/NanumGothic.ttf",
            "/System/Library/Fonts/AppleGothic.ttf"
        ]
        for alt_font in alt_fonts:
            if os.path.exists(alt_font):
                pdfmetrics.registerFont(TTFont('Korean', alt_font))
                break
        else:
            # 기본 폰트 사용
            pass
except Exception:
    # 폰트 등록 실패 시 기본 폰트 사용
    pass

class EvaluationExporter:
    """평가 데이터 추출 클래스"""
    
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self._setup_korean_styles()
    
    def _setup_korean_styles(self):
        """한글 지원 스타일 설정"""
        try:
            # 한글 제목 스타일
            self.styles.add(ParagraphStyle(
                name='KoreanTitle',
                parent=self.styles['Title'],
                fontName='Korean' if 'Korean' in [f.fontName for f in pdfmetrics.getRegisteredFontNames()] else 'Helvetica-Bold',
                fontSize=16,
                alignment=TA_CENTER,
                spaceAfter=20
            ))
            
            # 한글 본문 스타일
            self.styles.add(ParagraphStyle(
                name='KoreanNormal',
                parent=self.styles['Normal'],
                fontName='Korean' if 'Korean' in [f.fontName for f in pdfmetrics.getRegisteredFontNames()] else 'Helvetica',
                fontSize=10,
                alignment=TA_LEFT
            ))
            
            # 한글 소제목 스타일
            self.styles.add(ParagraphStyle(
                name='KoreanHeading',
                parent=self.styles['Heading2'],
                fontName='Korean' if 'Korean' in [f.fontName for f in pdfmetrics.getRegisteredFontNames()] else 'Helvetica-Bold',
                fontSize=12,
                alignment=TA_LEFT,
                spaceAfter=10
            ))
        except Exception:
            # 스타일 설정 실패 시 기본 스타일 사용
            pass
    
    def generate_filename(self, project_name: str, company_name: str, submitted_date: str, format_type: str) -> str:
        """파일명 생성 (프로젝트명_평가대상자명_평가일자.확장자)"""
        # 파일명에 사용할 수 없는 문자 제거
        safe_project = "".join(c for c in project_name if c.isalnum() or c in (' ', '-', '_')).strip()
        safe_company = "".join(c for c in company_name if c.isalnum() or c in (' ', '-', '_')).strip()
        
        # 날짜 형식 변환 (YYYY-MM-DD)
        try:
            if isinstance(submitted_date, str):
                date_obj = datetime.fromisoformat(submitted_date.replace('Z', '+00:00'))
            else:
                date_obj = submitted_date
            date_str = date_obj.strftime('%Y%m%d')
        except:
            date_str = datetime.now().strftime('%Y%m%d')
        
        extension = 'pdf' if format_type.lower() == 'pdf' else 'xlsx'
        filename = f"{safe_project}_{safe_company}_{date_str}.{extension}"
        
        return filename
    
    async def export_single_evaluation_pdf(self, evaluation_data: Dict[str, Any]) -> io.BytesIO:
        """단일 평가 데이터를 PDF로 추출"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*inch, bottomMargin=1*inch)
        
        story = []
        
        # 제목
        title = Paragraph(f"종합평가서", self.styles.get('KoreanTitle', self.styles['Title']))
        story.append(title)
        story.append(Spacer(1, 20))
        
        # 기본 정보 테이블
        basic_info_data = [
            ['프로젝트명', evaluation_data['project']['name']],
            ['평가대상기업', evaluation_data['company']['name']],
            ['담당자', evaluation_data['company'].get('contact_person', '')],
            ['연락처', evaluation_data['company'].get('phone', '')],
            ['평가표', evaluation_data['template']['name']],
            ['평가위원', evaluation_data['evaluator']['user_name']],
            ['제출일시', evaluation_data['sheet']['submitted_at'].strftime('%Y-%m-%d %H:%M') if evaluation_data['sheet'].get('submitted_at') else ''],
            ['총점', f"{evaluation_data['sheet'].get('weighted_score', 0):.1f}점"]
        ]
        
        basic_info_table = Table(basic_info_data, colWidths=[2*inch, 4*inch])
        basic_info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Korean' if 'Korean' in [f.fontName for f in pdfmetrics.getRegisteredFontNames()] else 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('BACKGROUND', (1, 0), (1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(basic_info_table)
        story.append(Spacer(1, 20))
        
        # 평가 항목별 상세 점수
        story.append(Paragraph("평가 항목별 상세", self.styles.get('KoreanHeading', self.styles['Heading2'])))
        story.append(Spacer(1, 10))
        
        # 평가 항목 테이블 헤더
        score_data = [['항목명', '설명', '배점', '획득점수', '가중치', '가중점수', '평가의견']]
        
        total_weighted_score = 0
        total_weight = 0
        
        for item in evaluation_data['template']['items']:
            score_info = next((s for s in evaluation_data['scores'] if s['item_id'] == item['id']), None)
            if score_info:
                weighted_score = score_info['score'] * item['weight']
                total_weighted_score += weighted_score
                total_weight += item['weight']
                
                score_data.append([
                    item['name'],
                    item['description'][:50] + '...' if len(item['description']) > 50 else item['description'],
                    f"{item['max_score']}점",
                    f"{score_info['score']}점",
                    f"{item['weight']}",
                    f"{weighted_score:.1f}점",
                    score_info['opinion'][:100] + '...' if len(score_info['opinion']) > 100 else score_info['opinion']
                ])
        
        # 최종 점수 행 추가
        final_average = total_weighted_score / total_weight if total_weight > 0 else 0
        score_data.append([
            '최종 점수', '', '', '', f"{total_weight}", f"{final_average:.1f}점", ''
        ])
        
        score_table = Table(score_data, colWidths=[1.2*inch, 1.5*inch, 0.7*inch, 0.8*inch, 0.6*inch, 0.8*inch, 1.4*inch])
        score_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), 'Korean' if 'Korean' in [f.fontName for f in pdfmetrics.getRegisteredFontNames()] else 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightblue),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'TOP')
        ]))
        
        story.append(score_table)
        
        # PDF 생성
        doc.build(story)
        buffer.seek(0)
        return buffer
    
    async def export_single_evaluation_excel(self, evaluation_data: Dict[str, Any]) -> io.BytesIO:
        """단일 평가 데이터를 Excel로 추출"""
        buffer = io.BytesIO()
        
        # openpyxl을 사용한 고급 Excel 생성
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "종합평가서"
        
        # 스타일 정의
        header_font = Font(bold=True, size=12)
        title_font = Font(bold=True, size=16)
        normal_font = Font(size=10)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 제목
        ws.merge_cells('A1:H1')
        ws['A1'] = "종합평가서"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # 기본 정보
        current_row = 3
        basic_info = [
            ('프로젝트명', evaluation_data['project']['name']),
            ('평가대상기업', evaluation_data['company']['name']),
            ('담당자', evaluation_data['company'].get('contact_person', '')),
            ('연락처', evaluation_data['company'].get('phone', '')),
            ('평가표', evaluation_data['template']['name']),
            ('평가위원', evaluation_data['evaluator']['user_name']),
            ('제출일시', evaluation_data['sheet']['submitted_at'].strftime('%Y-%m-%d %H:%M') if evaluation_data['sheet'].get('submitted_at') else ''),
            ('총점', f"{evaluation_data['sheet'].get('weighted_score', 0):.1f}점")
        ]
        
        for label, value in basic_info:
            ws[f'A{current_row}'] = label
            ws[f'B{current_row}'] = value
            ws[f'A{current_row}'].font = header_font
            ws[f'A{current_row}'].fill = header_fill
            ws[f'A{current_row}'].border = border
            ws[f'B{current_row}'].border = border
            current_row += 1
        
        # 평가 항목 헤더
        current_row += 2
        headers = ['항목명', '설명', '배점', '획득점수', '가중치', '가중점수', '평가의견']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        current_row += 1
        
        # 평가 항목 데이터
        total_weighted_score = 0
        total_weight = 0
        
        for item in evaluation_data['template']['items']:
            score_info = next((s for s in evaluation_data['scores'] if s['item_id'] == item['id']), None)
            if score_info:
                weighted_score = score_info['score'] * item['weight']
                total_weighted_score += weighted_score
                total_weight += item['weight']
                
                row_data = [
                    item['name'],
                    item['description'],
                    item['max_score'],
                    score_info['score'],
                    item['weight'],
                    round(weighted_score, 1),
                    score_info['opinion']
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=current_row, column=col)
                    cell.value = value
                    cell.border = border
                    if col in [3, 4, 6]:  # 숫자 컬럼 중앙 정렬
                        cell.alignment = Alignment(horizontal='center')
                
                current_row += 1
        
        # 최종 점수 행
        final_average = total_weighted_score / total_weight if total_weight > 0 else 0
        final_row_data = ['최종 점수', '', '', '', total_weight, round(final_average, 1), '']
        
        for col, value in enumerate(final_row_data, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = value
            cell.font = header_font
            cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
            cell.border = border
            if col in [3, 4, 6]:
                cell.alignment = Alignment(horizontal='center')
        
        # 컬럼 너비 조정
        column_widths = [15, 30, 8, 10, 8, 10, 40]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width
        
        # Excel 파일 저장
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    async def export_bulk_evaluations_excel(self, evaluations_data: List[Dict[str, Any]]) -> io.BytesIO:
        """여러 평가 데이터를 하나의 Excel 파일로 추출 (시트별 분리)"""
        buffer = io.BytesIO()
        
        with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
            # 요약 시트 생성
            summary_data = []
            for eval_data in evaluations_data:
                summary_data.append({
                    '프로젝트명': eval_data['project']['name'],
                    '평가대상기업': eval_data['company']['name'],
                    '평가표': eval_data['template']['name'],
                    '평가위원': eval_data['evaluator']['user_name'],
                    '제출일시': eval_data['sheet']['submitted_at'].strftime('%Y-%m-%d %H:%M') if eval_data['sheet'].get('submitted_at') else '',
                    '총점': round(eval_data['sheet'].get('weighted_score', 0), 1),
                    '상태': '제출완료' if eval_data['sheet']['status'] == 'submitted' else '평가중'
                })
            
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='전체요약', index=False)
            
            # 각 평가별 상세 시트 생성
            for i, eval_data in enumerate(evaluations_data):
                sheet_name = f"평가_{i+1}_{eval_data['company']['name'][:10]}"
                
                # 기본 정보
                basic_info = pd.DataFrame([
                    ['프로젝트명', eval_data['project']['name']],
                    ['평가대상기업', eval_data['company']['name']],
                    ['평가표', eval_data['template']['name']],
                    ['평가위원', eval_data['evaluator']['user_name']],
                    ['제출일시', eval_data['sheet']['submitted_at'].strftime('%Y-%m-%d %H:%M') if eval_data['sheet'].get('submitted_at') else ''],
                    ['총점', f"{eval_data['sheet'].get('weighted_score', 0):.1f}점"]
                ], columns=['항목', '내용'])
                
                # 평가 상세
                detail_data = []
                for item in eval_data['template']['items']:
                    score_info = next((s for s in eval_data['scores'] if s['item_id'] == item['id']), None)
                    if score_info:
                        weighted_score = score_info['score'] * item['weight']
                        detail_data.append({
                            '항목명': item['name'],
                            '설명': item['description'],
                            '배점': item['max_score'],
                            '획득점수': score_info['score'],
                            '가중치': item['weight'],
                            '가중점수': round(weighted_score, 1),
                            '평가의견': score_info['opinion']
                        })
                
                detail_df = pd.DataFrame(detail_data)
                
                # 시트에 데이터 쓰기
                startrow = 0
                basic_info.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False)
                startrow += len(basic_info) + 3
                detail_df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False)
        
        buffer.seek(0)
        return buffer
    
    def create_evaluation_summary_excel(self, evaluations_data: List[Dict[str, Any]]) -> io.BytesIO:
        """평가 요약 데이터를 Excel로 생성"""
        buffer = io.BytesIO()
        
        # 요약 데이터 준비
        summary_data = []
        for eval_data in evaluations_data:
            row = {
                'NO': len(summary_data) + 1,
                '프로젝트명': eval_data['project']['name'],
                '평가대상기업': eval_data['company']['name'],
                '담당자': eval_data['company'].get('contact_person', ''),
                '연락처': eval_data['company'].get('phone', ''),
                '평가표': eval_data['template']['name'],
                '평가위원': eval_data['evaluator']['user_name'],
                '제출일시': eval_data['sheet']['submitted_at'].strftime('%Y-%m-%d %H:%M') if eval_data['sheet'].get('submitted_at') else '',
                '총점': round(eval_data['sheet'].get('weighted_score', 0), 1),
                '상태': '제출완료' if eval_data['sheet']['status'] == 'submitted' else '평가중'
            }
            
            # 각 평가 항목별 점수 추가
            for item in eval_data['template']['items']:
                score_info = next((s for s in eval_data['scores'] if s['item_id'] == item['id']), None)
                if score_info:
                    row[f"{item['name']}_점수"] = score_info['score']
                    row[f"{item['name']}_의견"] = score_info['opinion'][:100] + '...' if len(score_info['opinion']) > 100 else score_info['opinion']
            
            summary_data.append(row)
        
        df = pd.DataFrame(summary_data)
        df.to_excel(buffer, index=False, sheet_name='평가결과요약')
        buffer.seek(0)
        return buffer

# 싱글톤 인스턴스
exporter = EvaluationExporter()
